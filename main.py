import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def log(msg):
    print(f"[LOG] {msg}")

def compare_tabs(filename):
    # Step 1: Read all sheets
    excel = pd.ExcelFile(filename, engine='openpyxl')
    sheet_names = excel.sheet_names
    log(f"Sheets found: {sheet_names}")

    # Step 2: Load baseline (first tab)
    baseline = pd.read_excel(filename, sheet_name=sheet_names[0], dtype=str)
    baseline = baseline.fillna("")
    log(f"Baseline: {sheet_names[0]} with {len(baseline)} rows")

    # Step 3: Prepare output dataframe
    output = pd.DataFrame()

    # Columns to compare/insert (hardcoded for now)
    compare_cols = ['A', 'D', 'F']  # compare values
    insert_cols = ['B', 'E']        # just copy

    # Step 4: Process each sheet
    for sheet in sheet_names:
        df = pd.read_excel(filename, sheet_name=sheet, dtype=str).fillna("")
        df['SourceSheet'] = sheet

        # Add insert columns (B, E)
        for col in insert_cols:
            if col in df.columns:
                output[col] = df[col]

        # Add compare columns (A, D, F)
        for col in compare_cols:
            if col in df.columns:
                output[f"{col}_{sheet}"] = df[col]

    # Step 5: Save output first (no highlight yet)
    temp_file = "combined_output.xlsx"
    output.to_excel(temp_file, index=False, engine='openpyxl')

    # Step 6: Apply yellow highlight where value != baseline
    wb = load_workbook(temp_file)
    ws = wb.active

    # Prepare fill style
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Find columns to compare (skip B, E)
    headers = [cell.value for cell in ws[1]]

    # Baseline sheet name
    baseline_name = sheet_names[0]

    for row in range(2, ws.max_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            if "_" in header and baseline_name not in header:
                base_col = header.split("_")[0] + "_" + baseline_name
                if base_col in headers:
                    base_idx = headers.index(base_col) + 1
                    val = ws.cell(row=row, column=col_idx).value
                    base_val = ws.cell(row=row, column=base_idx).value
                    if val != base_val:
                        ws.cell(row=row, column=col_idx).fill = yellow_fill

    wb.save(temp_file)
    log(f"Comparison complete. Output saved as {temp_file}")

def main():
    log("=== Compare Tabs Started ===")
    filename = input("Enter Excel file name (with multiple tabs): ").strip()
    compare_tabs(filename)
    log("=== Process Complete ===")

if __name__ == "__main__":
    main()
